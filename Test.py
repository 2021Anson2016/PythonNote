
# coding: utf-8

# In[8]:


from openpyxl import load_workbook
Destop = 'C:\\Users\\AnsonHsu\\Desktop\\'
wb2 = load_workbook( Destop + 'TC8100.xlsx')
type(wb2)
# wb2.get_sheet_names() 舊版
print(wb2.sheetnames) # sheetnames Returns the list of the names of worksheets in this workbook




sheetnames = wb2.sheetnames
# type(sheetnames)
# print(sheet.index)
sheet = wb2.active
rows = sheet.rows
# sheet.columns
for sheet in wb2:
    print(sheet.title)
    
    number_of_rows = sheet.rows
#     print(number_of_rows)
    number_of_columns = sheet.columns

    items = []

    rows = []  
    for row in range(1, number_of_rows):
        values = []
    for col in range(number_of_columns):
        value  = (sheet.cell(row,col).value)
        print(value)





# target = wb2.copy_worksheet(source)
sheet_shotline = wb2['shotline']
print(sheet_shotline)





from openpyxl import Workbook
import time

# book = Workbook()
sheet = wb2.active
print(sheet.cell)
# sheet.cell
# cells = sheet['A1': 'A9']
# print(cells)
# for c1, c2 in cells:
#     print("{0:8} {1:8}".format(str(c1.value), \
#                                str(c2.value)))


# In[ ]:


# 測試用
# sheet['A1'] = 56
# sheet['A2'] = 43
# now = time.strftime("%x")
# sheet['A3'] = now

# book.save(Destop + "sample.xlsx")


# In[3]:





# In[24]:


def add_agrs(arg1, arg2):
    print(arg1+arg2)
type(add_agrs)


# In[25]:


def run_something_with_args(func, arg1, arg2):  # 第一引數：傳入函數；第二、第三引數：為數值
    func(arg1, arg2)
run_something_with_args(add_agrs, 5, 3)


# In[26]:


def print_args(*args):   # *args：任意數量的引數
    print('Positional argument tuple：', args)
print_args()


# In[29]:


print_args(' ',3,2, 'asd', 'uha', ...)


# In[37]:


def  print_kwargs(**kwargs):
    print('keywird arguments:', kwargs)
print_kwargs(dessert='cake',people='anson' )


# In[38]:


help(max)


# In[39]:


help(max.__doc__)


# In[40]:


def run_with_positional_args(func, *args):
    return func(a)
    

